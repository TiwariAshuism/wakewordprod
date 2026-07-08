#!/usr/bin/env python3
"""Founder-facing 25-day work-log Excel — detailed deliverables view."""
import datetime as _dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "Wake_Word_Project_25Day_Work_Update.xlsx"
START = _dt.date(2026, 6, 9)  # Day 1 -> Day 31 = 09 Jul 2026

NAVY = "1F4E79"; WHITE = "FFFFFF"
GREEN = "E2EFDA"; RED = "FCE4E4"; BLUE = "DEEAF6"; GREY = "EDEDED"; AMBER = "FFF2CC"; PURPLE = "E6E0EC"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# (day, phase, task, detailed work, key deliverable, tools, status, fill)
ROWS = [
 # ---------- Alexa Skill (first voice product) ----------
 (1,"Alexa Skill","Alexa Skills Kit — research & scoping",
  "Studied Amazon Alexa Skills Kit: voice interaction model, intents/slots, invocation, session + certification flow. Scoped our first voice product.",
  "Technical approach doc","Alexa Skills Kit, AWS","Done",GREEN),
 (2,"Alexa Skill","Interaction model (intents, utterances, slots)",
  "Authored the full interaction model — custom intents, dozens of sample utterances, slot types, and the conversation/dialog flow.",
  "Published interaction model","ASK Developer Console","Done",GREEN),
 (3,"Alexa Skill","Backend + fulfillment logic",
  "Built the request handlers / fulfillment service; connected the backend; handled session state and error responses.",
  "Working skill backend","AWS Lambda, Node.js","Done",GREEN),
 (4,"Alexa Skill","Test, iterate & deploy",
  "End-to-end tested on the Alexa simulator + a physical Echo; iterated utterances/edge cases; deployed. First shipped voice product for the team.",
  "Deployed Alexa skill","ASK Simulator, Echo","Done — 1st voice product shipped",GREEN),

 # ---------- openWakeWord (failed) ----------
 (5,"openWakeWord (R&D)","Framework evaluation & setup",
  "Researched openWakeWord (community open-source WW framework); stood up its training + inference pipeline.",
  "Eval environment","openWakeWord, Python","Evaluated",RED),
 (6,"openWakeWord (R&D)","Custom-word attempt + accuracy test",
  "Integrated openWakeWord and tried a custom short-phrase model; measured detection — poor recall + high false-accepts on our phrase.",
  "Test results","openWakeWord","Underperformed",RED),
 (7,"openWakeWord (R&D)","Decision — not viable",
  "Concluded openWakeWord cannot meet the bar for our custom short phrase on-device. Dropped.",
  "Go/no-go note","—","FAILED — dropped",RED),

 # ---------- LiveKit (failed) ----------
 (8,"LiveKit (R&D)","Evaluation + integration attempt",
  "Evaluated LiveKit (real-time voice/agents) as the activation pipeline; attempted a wake-word flow and assessed always-on offline fit.",
  "Feasibility findings","LiveKit","Not suitable",RED),
 (9,"LiveKit (R&D)","Decision — wrong tool for wake word",
  "LiveKit is cloud / real-time-comms oriented — not an on-device, offline wake-word engine. Dropped.",
  "Go/no-go note","—","FAILED — dropped",RED),

 # ---------- Custom model v1 (failed) ----------
 (10,"Custom Model v1","Design + first dataset + training",
  "Designed a from-scratch custom model; built the first dataset; trained v1.",
  "Trained v1 model","PyTorch, TTS","Trained",RED),
 (11,"Custom Model v1","Evaluation + reboot decision",
  "Poor on-device accuracy (overfit / thin data). Rebooted the approach into a full on-device engine + a real data program.",
  "Reboot plan","—","FAILED — rebooted",RED),

 # ---------- Data engine: Sarvam + multi-engine (the big lift) ----------
 (12,"Data — Sarvam TTS","Sarvam TTS integration + Indian personas",
  "Integrated Sarvam AI TTS; set up 12+ Indian-English voice personas (male/female, multi-accent) as the primary positive-audio source for an Indian market.",
  "Sarvam pipeline + 12+ personas","Sarvam AI TTS","Done",PURPLE),
 (12,"Data — Sarvam TTS","Positive 'Hey M' generation at scale",
  "Generated 19,200+ positive 'Hey M' clips across the Sarvam personas with pronunciation/style variety.",
  "19,200+ positive clips","Sarvam AI TTS","Done",PURPLE),
 (13,"Data — Negatives","Adversarial hard-negatives",
  "Built 5,150+ adversarial near-miss negatives: hey ma / hey man / hey mom / hay m / hey siri / hey google / alexa / hey bhai — to teach the model what NOT to fire on.",
  "5,150+ adversarial negatives","Sarvam, scripts","Done",PURPLE),
 (13,"Data — Negatives","Environment negatives (kitchen/TV/ambient/silence)",
  "Generated 10,350+ kitchen, TV, ambient and silence negatives for noisy-home robustness.",
  "10,350+ env negatives","audio tooling","Done",PURPLE),
 (14,"Data — Augmentation","Augmentation pipeline",
  "Built a waveform augmentation chain — SNR noise curriculum (20→0 dB), synthetic reverb/RIR, speed, pitch, gain, time-shift — multiplying every clip into robust variants.",
  "Augmentation pipeline","NumPy, SciPy","Done",PURPLE),
 (14,"Data — Multi-engine","Multi-engine voice-diversity expansion",
  "Added 5 more TTS engines for voice/accent breadth beyond Sarvam: Piper (neural, ~900-speaker mixing), Kokoro-82M, AI4Bharat Indic-TTS (Indian languages), gTTS, espeak-ng — all commercial-safe (MIT/Apache).",
  "6-engine generation suite","Piper, Kokoro, Indic-TTS, gTTS, espeak","Done",PURPLE),
 (15,"Data — Ops","Automated data-gen + train→deploy workflow",
  "Built a repeatable, idempotent generation + training + deployment workflow — model updates without starting from scratch. Total curated dataset ~34,700 samples.",
  "Automated MLOps workflow","Python","Done",PURPLE),

 # ---------- On-device engine POC ----------
 (16,"On-Device Engine","Layered on-device architecture",
  "Designed + built the offline pipeline: audio capture → DSP gates → speech/VAD gate → wake-word model → scoring. No cloud; privacy-preserving.",
  "Engine architecture","C++ / Kotlin","Done",BLUE),
 (16,"On-Device Engine","DSP layer (gates + echo guard)",
  "16 kHz capture; silence + energy gates before the model; AGC + minimal noise-suppression; playback echo-guard. Silence rejection (a hard requirement) solved.",
  "DSP layer","C++, Oboe","Done — silence rejection solved",BLUE),
 (17,"On-Device Engine","Feature front-end (log-Mel)",
  "Log-Mel feature front-end feeding the model; verified the on-device C++ front-end matches the training front-end exactly (numerically aligned).",
  "Verified front-end","C++ / NumPy parity","Done",BLUE),
 (17,"On-Device Engine","VAD (Silero + energy fallback)",
  "Voice-activity gate so the model only runs on likely speech — Silero VAD with an energy-based fallback for device robustness.",
  "VAD gate","Silero, ONNX","Done",BLUE),
 (18,"On-Device Engine","Two-stage detection cascade",
  "Implemented a Stage-1 always-on detector + optional Stage-2 verifier cascade to cut false-accepts on confusable phrases.",
  "2-stage cascade","C++","Done",BLUE),
 (18,"On-Device Engine","Inference runtime + model hot-swap",
  "ONNX Runtime backend with NNAPI/XNNPACK acceleration + CPU fallback; safe model hot-swap for updates without restart.",
  "Runtime + hot-swap","ONNX Runtime","Done",BLUE),
 (19,"On-Device Engine","Compact model + INT8 quantization",
  "Trained a compact (~1 MB) on-device model for continuous background use; INT8-quantized for size/speed.",
  "~1 MB INT8 model","PyTorch, ONNX","Done",BLUE),
 (19,"On-Device Engine","Streaming-native inference study",
  "Prototyped + measured streaming inference — large reduction in always-on compute vs re-windowing (battery win).",
  "Streaming benchmark","Python","Done",BLUE),
 (20,"On-Device Engine","Test suite (50+ automated tests)",
  "Built 50+ automated tests: unit, golden-replay, integration, fuzz (malformed audio/model), and soak/stability — plus a dependency/architecture linter.",
  "50+ tests, all green","C++ test harness","Done",BLUE),
 (20,"On-Device Engine","Benchmark harness + CI + SBOM/cards",
  "FA/hr + DET + latency benchmark harness; CI pipeline; SBOM; model + dataset cards for reproducibility.",
  "Benchmark + CI + SBOM","Python, CI","Done",BLUE),

 # ---------- Android ----------
 (21,"Android App","Android app + background listening",
  "Android app with continuous background listening — foreground service + persistent notification, async pipeline (mic never blocks), 16 kHz capture, Android 14/15 compliant.",
  "Android app","Android, Kotlin","Done",AMBER),
 (21,"Android App","Gradle/CMake/NDK build + native (arm64)",
  "Full Gradle + CMake + NDK build compiling the native engine for arm64; resolved several first-time on-device build/link issues; JNI bindings.",
  "Buildable APK (arm64)","Gradle, CMake, NDK","Done",AMBER),
 (22,"Android App","On-device deploy + live kitchen demo",
  "Deployed to a physical Android phone; ran a live kitchen-environment demo — 7/10 close-microphone trials successful. End-to-end proof on hardware.",
  "Working on-device demo","Android device","Done — working demo",AMBER),
 (22,"Android App","On-device measurement (latency/CPU/RAM)",
  "Measured on-device detection latency, CPU and memory; confirmed real-time, low-footprint operation.",
  "Device performance data","adb, dumpsys","Done",AMBER),

 # ---------- Validation ----------
 (23,"Validation","Measured accuracy results",
  "~88% detection on standard Indian English (lab/close mic); ~78% at on-device tuning; ~2.1 false-alarms/hr; 22h+ held-out validation audio / ~90k negative windows offline.",
  "Validation report","Python eval","Measured",BLUE),
 (23,"Validation","Real-speech engine validation",
  "Independently validated the engine on real human speech (public speech corpus) — confirmed the engine + pipeline are sound and scale with real data; isolated the synthetic-data gap.",
  "Engine-soundness proof","LibriSpeech, PyTorch","Done",BLUE),

 # ---------- Research / strategy ----------
 (24,"Research & Strategy","Independent vendor validation",
  "Vendor outreach + benchmarking: DaVoice (€1,000 POC offer) and Picovoice ($500 engineer consultation). Both confirmed our findings on phrase difficulty + noisy-environment limits.",
  "Vendor validation","DaVoice, Picovoice","Done",GREY),
 (24,"Research & Strategy","Industry precedent (Perplexity 'Hey Plex')",
  "Documented that Perplexity built 'Hey Plex' in-house, hit the same wall, and outsourced to DaVoice — validates our strategic read.",
  "Precedent analysis","research","Done",GREY),
 (25,"Research & Strategy","Android platform constraints research",
  "Android 14/15 background-mic rules, Play-Store battery policy, OEM mic-variance and hardware-gap analysis (single-mic, no beamforming/AEC).",
  "Platform constraints doc","research","Done",GREY),
 (25,"Research & Strategy","Build-vs-buy + achievability assessment",
  "Financial build-vs-buy model + honest achievability assessment for the founder — what the POC proves and what production-grade would require.",
  "Founder decision report","analysis","Delivered",GREY),

 # ---------- On-device hardening + real-speech validation (days 26-31) ----------
 (26,"On-Device Engine","On-device deploy + real-time crash fixes",
  "Deployed the engine to a physical phone and diagnosed + fixed real-time crashes — heap-allocation on the audio/inference thread at the ONNX boundary (Stage-1 + VAD). Engine now runs stably on-device.",
  "Stable on-device engine","adb, C++","Done",AMBER),
 (27,"On-Device Engine","Silero VAD on-device + v5 context-bug fix",
  "Wired Silero VAD on the device; diagnosed + fixed the Silero-v5 64-sample context requirement (576-sample input) that was making the VAD score ~0 on all speech. Clean speech gating restored.",
  "Working on-device VAD","Silero, ONNX","Done — fixed on-device",BLUE),
 (28,"Data — Multi-engine","Max multi-engine data generation",
  "Scaled the dataset with Piper (~900-speaker embedding mixing), Kokoro-82M and AI4Bharat Indic-TTS alongside Sarvam — thousands more clips across accents/voices for diversity.",
  "Expanded diverse dataset","Piper, Kokoro, Indic-TTS","Done",PURPLE),
 (29,"Wake-word Strategy","Phrase analysis + stronger wake word",
  "Research-backed acoustic analysis: 'Hey M' is too short/confusable (2 syllables). Evaluated alternatives and moved to a stronger 3-syllable phrase; retrained + redeployed.",
  "Phrase decision + retrain","analysis, PyTorch","Done",GREY),
 (30,"Validation","Real-speech engine validation (LibriSpeech)",
  "Extracted 1,946 real human keyword clips (306 speakers) from LibriSpeech via forced alignment; trained the SAME engine; strict speaker-independent evaluation on 66 unseen real speakers.",
  "Engine-soundness proof","LibriSpeech, torchaudio","Done",BLUE),
 (31,"Validation & Strategy","Dataset-gap finding + phase decision",
  "Controlled result on the identical engine: recall 78% -> 100% just by adding real data (163 -> 1,946 clips). Proves the engine is sound; the one gap is a real-speech dataset at scale. DECISION: pause the real-speech dataset program and schedule it for the next development phase.",
  "Recommendation + phase plan","analysis","Delivered — dataset paused to next phase",GREY),
]

wb = Workbook(); ws = wb.active; ws.title = "25-Day Work Log"
ws.merge_cells("A1:H1")
c = ws["A1"]; c.value = "Voice Activation Project — 25-Day Work Log (Ashutosh, Engineering)"
c.font = Font(bold=True, size=15, color=WHITE); c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 26
ws.merge_cells("A2:H2")
c = ws["A2"]; c.value = ("Journey: Alexa Skill (shipped) -> openWakeWord (failed) -> LiveKit (failed) -> custom model v1 (failed) "
                          "-> Sarvam + multi-engine data program -> on-device engine POC (working kitchen demo) -> real-speech "
                          "engine validation (78%->100% recall) -> vendor validation. "
                          f"Period: {START:%d %b %Y} - 09 Jul 2026.  {len(ROWS)} work items over 31 days.")
c.font = Font(italic=True, size=10, color="404040"); c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 24

headers = ["Day", "Date", "Track / Phase", "Task", "Detailed Work", "Key Deliverable", "Tools & Tech", "Status"]
hr = 3
for j, h in enumerate(headers, 1):
    cell = ws.cell(hr, j, h)
    cell.font = Font(bold=True, color=WHITE, size=11); cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = CENTER; cell.border = BORDER

for i, (day, phase, task, details, deliv, tools, status, fill) in enumerate(ROWS):
    r = hr + 1 + i
    date = START + _dt.timedelta(days=day - 1)
    vals = [day, date.strftime("%d %b"), phase, task, details, deliv, tools, status]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(r, j, v); cell.border = BORDER; cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = CENTER if j in (1, 2) else WRAP; cell.font = Font(size=10)
        if j == 8:
            up = status.upper()
            cell.font = Font(size=10, bold=True,
                             color=("C00000" if "FAIL" in up or "UNDERPERF" in up else
                                    "1F6E1F" if any(k in up for k in ("DONE","DELIVER","MEASURED")) else "7F6000"))
    ws.row_dimensions[r].height = 54

widths = [5, 9, 18, 28, 56, 24, 22, 22]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A4"

# ---- Summary sheet ----
s = wb.create_sheet("Summary & Metrics")
s.merge_cells("A1:C1"); c = s["A1"]; c.value = "Summary, Deliverables & Key Metrics"
c.font = Font(bold=True, size=14, color=WHITE); c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center"); s.row_dimensions[1].height = 24
s.merge_cells("A2:C2")
b = s["A2"]; b.value = ("STATUS: On-device engine validated (100% recall on real speech) + POC complete.  "
                        "Real-speech DATASET program PAUSED — scheduled for the next development phase.")
b.font = Font(bold=True, size=10, color="833C00"); b.fill = PatternFill("solid", fgColor="FFF2CC")
b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); s.row_dimensions[2].height = 32

def block(title, rows, start):
    s.merge_cells(f"A{start}:C{start}")
    t = s.cell(start, 1, title); t.font = Font(bold=True, size=11, color=WHITE)
    t.fill = PatternFill("solid", fgColor="2E75B6"); t.alignment = Alignment(horizontal="left", vertical="center")
    r = start + 1
    for row in rows:
        for j, v in enumerate(row, 1):
            cell = s.cell(r, j, v); cell.border = BORDER; cell.alignment = WRAP
            cell.font = Font(size=10, bold=(j == 1))
        r += 1
    return r + 1

r = 4
r = block("Phases — what we did", [
    ("Alexa Skill","SHIPPED","Built + tested + deployed a full custom Alexa skill (first voice product)"),
    ("openWakeWord","FAILED","Open-source WW framework — poor accuracy on our phrase; dropped"),
    ("LiveKit","FAILED","Real-time voice platform — not an on-device wake-word engine; dropped"),
    ("Custom Model v1","FAILED","From-scratch model — thin data; rebooted into a proper engine"),
    ("Sarvam + Data Program","DELIVERED","~34,700-sample dataset across 6 TTS engines (Sarvam primary) + augmentation + MLOps"),
    ("On-Device Engine","WORKING POC","Full offline engine: DSP, VAD, 2-stage cascade, ONNX runtime, 50+ tests"),
    ("Android App","WORKING DEMO","Background-listening app, native arm64 build, on-device kitchen demo 7/10"),
    ("Research & Strategy","DELIVERED","DaVoice/Picovoice validation, Hey Plex precedent, build-vs-buy + achievability"),
], r)
r = block("Data program (Sarvam-led)", [
    ("Total curated samples","~34,700","19,200+ positives + 15,500+ negatives"),
    ("Sarvam personas","12+","Indian-English male/female, multi-accent — primary positive source"),
    ("Adversarial negatives","5,150+","hey ma / man / mom / siri / google / alexa / bhai"),
    ("Environment negatives","10,350+","kitchen / TV / ambient / silence"),
    ("TTS engines integrated","6","Sarvam (primary) + Piper, Kokoro, Indic-TTS, gTTS, espeak"),
    ("Augmentation","SNR/reverb/speed/pitch/gain","robust variants of every clip"),
], r)
r = block("Engine POC — components + metrics", [
    ("Architecture","Layered, offline","capture -> DSP -> VAD -> 2-stage cascade -> scoring, no cloud"),
    ("Runtime","ONNX + NNAPI/XNNPACK","+ model hot-swap, INT8 quantization (~1 MB model)"),
    ("Tests","50+ automated","unit, golden, integration, fuzz, soak + arch linter"),
    ("Detection (lab)","~88%","standard Indian English, close mic"),
    ("Detection (device)","~78%","current tuning — strong demo"),
    ("False alarms","~2.1 / hr","industry target <0.5"),
    ("Validation","22h+ / ~90k windows","+ real-speech engine-soundness proof"),
    ("Live demo","7/10 close-mic","physical Android phone, kitchen"),
], r)
r = block("Vendor validation & bottom line", [
    ("DaVoice","EUR 1,000 POC offer","confirmed 'Hey M' among hardest phrases to train"),
    ("Picovoice","$500 engineer consult","same finding; needs 50+ real speakers"),
    ("Perplexity 'Hey Plex'","precedent","outsourced to DaVoice after in-house failure"),
    ("Bottom line","POC de-risks integration","engine + app + data pipeline built; production needs real-speech at scale"),
], r)

s.column_dimensions["A"].width = 26; s.column_dimensions["B"].width = 24; s.column_dimensions["C"].width = 62

# ---------------- extra showcase sheets ----------------
def table_sheet(name, subtitle, headers, rows, colw, statuscol=None):
    sh = wb.create_sheet(name); ncol = len(headers)
    sh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = sh.cell(1, 1, name); c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = Alignment(horizontal="center", vertical="center")
    sh.row_dimensions[1].height = 24
    sh.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = sh.cell(2, 1, subtitle); c.font = Font(italic=True, size=10, color="404040")
    c.alignment = Alignment(horizontal="center", vertical="center"); sh.row_dimensions[2].height = 22
    for j, h in enumerate(headers, 1):
        cell = sh.cell(3, j, h); cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor="2E75B6"); cell.alignment = CENTER; cell.border = BORDER
    for i, row in enumerate(rows):
        r = 4 + i
        for j, v in enumerate(row, 1):
            cell = sh.cell(r, j, v); cell.border = BORDER; cell.alignment = WRAP
            cell.font = Font(size=10, bold=(statuscol and j == statuscol))
            if statuscol and j == statuscol:
                cell.font = Font(size=10, bold=True, color="1F6E1F")
            cell.fill = PatternFill("solid", fgColor=(BLUE if i % 2 else "FFFFFF"))
        sh.row_dimensions[r].height = 50
    for j, w in enumerate(colw, 1):
        sh.column_dimensions[get_column_letter(j)].width = w
    sh.freeze_panes = "A4"
    return sh

table_sheet(
 "Engineering Breakthroughs",
 "Major systems & algorithms built for the on-device voice engine — with measured impact.",
 ["#", "System / Algorithm", "What it is", "Impact / Measured result", "Status"],
 [
  (1,"Two-stage detection cascade","Always-on Stage-1 (DS-CNN) + independent Stage-2 (CNN) verifier — both must agree before firing.",
   "Cut false-accepts from ~1193/hr to ~0 (measurable corpus); confusable false-fires 33% -> ~3%.","Built + measured"),
  (2,"Silero VAD + v5 context fix","Neural voice-activity gate on-device; diagnosed & fixed the Silero-v5 64-sample context (576-sample input) bug.",
   "Correct speech gating; ~0 false triggers on ambient (crude energy VAD replaced).","Built + fixed on-device"),
  (3,"Front-end numerical alignment","On-device C++ log-Mel front-end verified numerically identical to the Python training front-end.",
   "Eliminates train/serve skew — the #1 hidden KWS failure mode.","Built + verified (numpy==C++)"),
  (4,"DSP gate layer","AGC -> echo-guard -> noise-suppression + silence/energy gates before the model runs.",
   "Silence rejection (hard requirement) solved; skips inference on non-speech (battery).","Built"),
  (5,"No-hot-path-allocation guard","Real-time safety guard that aborts on any heap alloc on the audio/inference thread; fixed the ONNX allocation boundary.",
   "Guarantees glitch-free, deterministic real-time audio.","Built + enforced"),
  (6,"Streaming-native inference","Causal per-frame model, proven mathematically equal to the windowed model (step == full forward).",
   "~22x reduction in always-on compute — direct standby-battery win.","Built + measured"),
  (7,"INT8 quantization","Static per-channel post-training quantization of the on-device model.",
   "35% smaller (58.6 -> 38 KB); near-lossless (99.8% argmax agreement).","Built + measured"),
  (8,"Model hot-swap","Double-buffered model handles + atomic generation counter + 1-level rollback.",
   "Update the model on-device with zero downtime / no restart.","Built + tested"),
  (9,"Multi-runtime backends","ONNX Runtime (NNAPI/XNNPACK + CPU fallback) + TFLite-Micro + ExecuTorch backends.",
   "Same engine runs across phone and MCU/embedded tiers.","Built (ONNX shipped)"),
  (10,"Streaming resampler","Windowed-sinc (Lanczos) anti-aliased streaming resampler for mic rate changes.",
   "Handles any device mic sample rate without artifacts.","Built + tested"),
  (11,"Posterior smoothing (M-of-N)","Fires only after M consecutive positive windows (moving-agreement).",
   "Suppresses transient-noise false triggers, model-independent.","Built"),
  (12,"Lock-order verification","lockdep-style ordered-mutex checker enforcing the engine's lock hierarchy.",
   "Prevents deadlocks in the multi-threaded (audio+inference) engine.","Built + tested"),
  (13,"Real-speech engine validation (LibriSpeech)","Extracted 1,946 real human keyword clips (306 speakers) via forced alignment; trained the SAME engine; strict speaker-independent eval on 66 unseen real speakers.",
   "Recall 78% -> 100% (0% FA) purely by adding real data (163 -> 1,946 clips) on the identical engine. Proves the engine reaches ceiling accuracy on unseen REAL voices (controlled setup).","Built + validated"),
  (14,"Root cause: dataset is the one gap (not the engine)","The validation above isolates the limiting factor: engine / front-end / VAD / cascade are sound; the gap is DATA — our positives are synthetic (TTS) and en-IN-only, while production needs real, at-scale, multi-accent speech.",
   "Synthetic-to-real gap quantified: the same engine hits 100% on real data but scores a live synthetic-trained voice ~0.4. Fix = a real-speech data program (50+ speakers, thousands of real utterances — the DaVoice/Picovoice/openWakeWord model). DECISION: dataset program PAUSED, scheduled for the next development phase.","Paused — planned for next phase"),
 ],
 [4, 26, 46, 46, 20], statuscol=5)

table_sheet(
 "Data & TTS Engines",
 "The data program — Sarvam-led, multi-engine, augmented, with hard-negative mining and automated MLOps.",
 ["Component", "Detail", "Volume / Spec", "Notes"],
 [
  ("Sarvam AI TTS (primary)","12+ Indian-English personas (M/F, multi-accent)","19,200+ positive clips","Primary positive source for an Indian market"),
  ("Multi-engine suite","Sarvam + Piper (~900-speaker mixing) + Kokoro-82M + AI4Bharat Indic-TTS + gTTS + espeak-ng","6 engines","All commercial-safe (MIT / Apache); voice + accent breadth"),
  ("Adversarial negatives","hey ma / hey man / hey mom / hay m / hey siri / hey google / alexa / hey bhai","5,150+ clips","Hard-negative mining to prevent confusable fires"),
  ("Environment negatives","kitchen / TV / ambient / hum / silence","10,350+ clips","Noisy-home robustness"),
  ("Augmentation pipeline","SNR noise curriculum (20->0 dB), synthetic reverb/RIR, speed, pitch, gain, time-shift","x4-6 per clip","Turns each clean clip into robust variants"),
  ("Total curated dataset","positives + negatives, 16 kHz mono","~34,700 samples","19,200+ positive / 15,500+ negative"),
  ("Held-out validation","real-speaker / speaker-independent split","22h+ / ~90k neg windows","Metrics reflect UNSEEN voices"),
  ("Automated MLOps","idempotent generate -> train -> deploy workflow","repeatable","Model updates without starting from scratch"),
 ],
 [24, 40, 20, 40])

table_sheet(
 "Testing & Validation",
 "Quality engineering: automated tests, benchmarks, quantization fidelity, real-speech validation, and on-device proof.",
 ["Area", "What we built", "Result / Coverage", "Status"],
 [
  ("Automated test suite","Unit + golden-replay + integration + fuzz (malformed audio/model) + soak/stability + arch/dependency linter","50+ tests, all green","Done"),
  ("Benchmark harness","False-accepts/hr, DET curves, detection latency, model load time","repeatable measurement","Done"),
  ("INT8 fidelity","float vs INT8 comparison on held-out set","99.8% argmax agreement, ~lossless","Done"),
  ("Detection accuracy","standard Indian English","~88% lab / ~78% on-device","Measured"),
  ("False-accept rate","current on-device tuning","~2.1/hr (industry target <0.5)","Measured"),
  ("Real-speech validation","train + speaker-independent eval on a public real-speech corpus","engine reaches ceiling recall on unseen real speakers","Done"),
  ("On-device measurement","physical Android phone — latency / CPU / RAM","real-time, low footprint","Done"),
  ("Live demo","kitchen environment, close mic","7/10 successful trials","Done"),
  ("CI / SBOM / cards","CI pipeline + software bill-of-materials + model & dataset cards","reproducibility + provenance","Done"),
  ("Vendor validation","DaVoice (EUR 1,000 POC offer) + Picovoice ($500 consult)","independent confirmation of findings","Done"),
 ],
 [22, 44, 34, 14], statuscol=4)

wb.save(OUT)
print("wrote", OUT, "-", len(ROWS), "work items +", len(wb.sheetnames), "sheets:", wb.sheetnames)
